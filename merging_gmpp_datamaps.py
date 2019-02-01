'''this is a programme to merge the old gmpp datamap (dm) with the latest one for the new gmpp template. this will
probably be some throw away code that is not required after this - as the new template will be used in the future

I stopped work on this as below gave me a viable dm to work on

programme is incomplete and there were some bugs which I couldn't explain'''

from openpyxl import load_workbook, Workbook
from collections import OrderedDict
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime
from bcompiler.utils import project_data_from_master

def merging_dms(old_dm, new_dm):
    wb = Workbook()
    ws = wb.active

    new_old_keys = new_dm['Old GMPP']  # reference to the old key values in the new datamap
    old_old_keys = old_dm['gmpp_template_cell_reference'] # reference to the old key values in the old datamap


    for i, key_1 in enumerate(new_old_keys):
        for key_2 in old_old_keys:
            if new_old_keys[key_1] == 'GMPP Return\'!' + old_old_keys[key_2]:
                ws.cell(row=i+2, column=4).value = key_2
                print(i, key_1)

    return wb

old = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\gmpp_reporting_docs\\datamaps\\gmpp_datamap_2.0.xlsx")

new = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\gmpp_reporting_docs\\datamaps\\new_gmpp_datamap.xlsx")

merged_wb = merging_dms(old, new)

merged_wb.save("C:\\Users\\Standalone\\Will\\masters folder\\gmpp_reporting_docs\\datamaps\\test_merge.xlsx")