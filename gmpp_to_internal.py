'''programme to move gmpp master data over to dft internal master data format '''

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master

def create_internal_master(gmpp_data, dft_data, dft_dm):
    ws = dft_dm.active

    for i, name in enumerate(gmpp_data):
        print(name)
        ws.cell(row=1, column=4+i).value = name  # place project names in file

        dm_list = []
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value
            dm_list.append(key)

        # for loop for placing data into the worksheet
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value
            # this loop places all latest raw data into the worksheet
            try:
                if key in gmpp_data[name].keys():
                    ws.cell(row=row_num, column=4+i).value = gmpp_data[name][key]
                else:
                    ws.cell(row=row_num, column=4 + i).value = dft_data[name][key]
            except KeyError:
                pass

    return dft_dm


dft_datamap = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\dms\\datamap_current_internal.xlsx")

gmpp_master = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\core data\\gmpp_master_3"
                                       "_2018.xlsx")
dft_master = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_3_2018.xlsx")

output = create_internal_master(gmpp_master, dft_master, dft_datamap)

output.save("C:\\Users\\Standalone\\Will\\gmpp_data_output.xlsx")