'''Programme to move internal master data into the same order as gmpp master data

Output - Excel master file with internal data ordered as per the gmpp master data. In addition internal data missing
from the gmpp data set is included at the bottom. project data in the output file should be cut and paste into the
gmpp master dataset.

input documents
dft_data = latest internal quarter master - for non-gmpp projects
gmpp_dm = Gmpp datamap. NOTE. difference variation from datamap that is passed into bcompiler. Difference is that
internal keys not included in gmpp template are included at the bottom of the file'''

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master

def create_internal_master(dft_data, gmpp_dm):
    ws = gmpp_dm.active

    for i, name in enumerate(dft_data):
        print(name)
        ws.cell(row=1, column=4+i).value = name  # place project names in file

        dm_list = []
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value
            dm_list.append(key)

        # for loop for placing data into the worksheet
        for row_num in range(2, ws.max_row+1):
            key = ws.cell(row=row_num, column=1).value
            # this loop places all latest raw data into the worksheet
            try:
                if key in dft_data[name].keys():
                    ws.cell(row=row_num, column=4+i).value = dft_data[name][key]
                else:
                    pass
                    #ws.cell(row=row_num, column=4 + i).value = dft_data[name][key]
            except KeyError:
                pass

    return gmpp_dm


gmpp_datamap = load_workbook("C:\\Users\\Standalone\\Will\\masters folder\\dms\\dm_merged_all_excel_master.xlsx")

dft_master = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\core data\\master_non_gmpp_testing.xlsx")

output = create_internal_master(dft_master, gmpp_datamap)

output.save("C:\\Users\\Standalone\\Will\\output_internal_data_gmpp_format.xlsx")