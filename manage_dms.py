'''Currently programme that has ability to check difference between two datamaps'''

from openpyxl import load_workbook, Workbook
from bcompiler.utils import project_data_from_master

def place_in_excel(list):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = 'MISSING'
    for i, key in enumerate(list):
        ws.cell(row=i+2, column=1).value = key

    return wb


gmpp_datamap = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\dms\\gmpp_dm_merged_excel_master.xlsx")

#internal_datamap = project_data_from_master("C:\\Users\\Standalone\\Will\\masters folder\\dms\\datamap_current_internal"
#                                            ".xlsx")

gmpp_keys = gmpp_datamap['sheet'].keys()

#internal_keys = internal_datamap['template_sheet'].keys()

#missing_keys = [x for x in internal_keys if x not in gmpp_keys]

#output = place_in_excel(missing_keys)

#output.save("C:\\Users\\Standalone\\Will\\test.xlsx")